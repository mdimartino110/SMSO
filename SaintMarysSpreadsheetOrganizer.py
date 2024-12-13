import openpyxl
import datetime
import tkinter
import sys
from tkinter import Tk     # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
from tkinter import filedialog
from tkinter import *
from tkinter import ttk

# Python Script to take in a exports of student information
# and format a new spreadsheet to submit to various sources
# such as New York State or MySchoolApp
# Version: 4.0
# Major Update: Added Manhasset District Upload Functionality
# Notes for future: 2025-2026 School Year will change to Earth and Space Science Regents.
# Course Code: 03008

today = datetime.date.today()
month = today.month
year = today.year
i = 0
#nameList = []
#foodList = []
#name = "Name"
#food = "Food"
# Defining a lot of stuff to use for later
longCode = "80049051"
longCodeList = ["Long Code"]
districtCodeList = ["District Code"] # Used with longCode for Regents Upload.
shortCode = "049051"
shortCodeList = ["Short"]
locationCodeList = ["LocationCode"]
regentDateList = ["Regent Date"]
adminMonthList = []
textMonth = ""
if month >=9:
    regentDate = str(year + 1) + "-06-30"
    textMonth = "Jan"
elif month <= 5:
    textMonth = "Jun"
else:
    regentDate = str(year) + "-06-30"
    textMonth = "Aug"
stuIDList = []
fNameList = []
lNameList = []
mNameList = []
gradeList = []
startList = ["Start"]
hrList = []
dobList = []
genderList = []
ethnicityList = []
zipCodeList = []
A = "A"
aList = []
#startDate
addressList = []
address2List = []
cityList = []
stateList = []
p1fnList = []
p1lnList = []
p2fnList = []
p2lnList = []
p1fullList = ["Contact 1"]
p2fullList = ["Contact 2"]
hispanicList = []
schoolList = ["School"]
school = "Saint Mary's High School"
spacerList = []
courseList = []
sectionList = []
teacherList = []
versionList = []
adminMonth = []
scoreList = []
junkList = []
hostIDList = []
userIDList = []
stuEmailList = []
titleList = ["District Code", "Location Code", "Version", "Admin Month", \
             "StudentID", "LastName", "FirstName", "GradeLevel", \
             "CourseSection", "TeacherName", "StateCourseCode", \
             "DistrictName", "School"]
userID = ["user_id"]
testDateTime = ["test_date_time"]
comment = ["comment"]
overallScore = ["overall_score"]
registeredDate = ["registered_date"]
registeredLocation = ["registered_location"]
rescheduled = ["rescheduled"]
noShow = ["no_show"]
proctorID = ["proctor_id"]
proctorOther = ["proctor_other"]
printOnReportCard = ["print_on_report_card"]
printOnTranscript = ["print_on_transcript"]
percentile = ["percentile"]
scale = ["scale"]
stanine = ["stanine"]
subtestComment = ["subtest_comment"]
p1emailList = []
p1cellList = []
p1workList = []
p1homeList = []
p2workList = []
I20List = []
stuEmailList = []
districtList = []

    
distList = ["Code"] # Used with district codes sheet for demographic information.
# This dictionary is used to organize the data taken in from the  demographic upload list. 
#Key is the first value in the column (title). Value is the list you add the data to.
dict1 = {
    "STU_ID":stuIDList,
    "FNAME":fNameList,
    "LNAME": lNameList,
    "MI": mNameList,
    "GRADE": gradeList,
    "ADD1": addressList,
    "ADD2": address2List,
    "CITY": cityList,
    "STATE": stateList,
    "ZIP": zipCodeList,
    "GEN": genderList,
    "DOB": dobList,
    "Parent 1 First Name":p1fnList,
    "Parent 1 Last Name":p1lnList,
    "Parent 2 First Name":p2fnList,
    "Parent 2 Last Name":p2lnList,
    "ETH":ethnicityList,
    "Hispanic?":hispanicList,
    "HR":hrList}

# This dictionary is for formatting ethnicity values.
dict2 = {
    "1":"W",
    "2":"B",
    "3":"A",
    "4":"W",
    "5":"I",
    "6":"W",
    "7":"P"}

#month = 1
if month <= 8 and month >= 12:
    year = year - 1
else:
    pass

dict3 = {
    9:year,
    10:year - 1,
    11:year - 2,
    12:year - 3}

dict4 = { # Regents Course Code
    "Algebra 1 Common Core":"02050",
    "Algebra 1H Common Core":"02050",
    "Algebra II":"02056CC",
    "Algebra II Honors":"02056CC",
    "Chemistry":"03101",
    "U.S. History":"04101F",
    "U.S. History H":"04101F",
    "AP U.S. History":"04101F",
    "Physics Honors":"03151",
    "Biology":"03051",
    "Biology H":"03051",
    "Global History 10":"04052NF",
    "Global History 10 H":"04052NF",
    "AP World History: Modern":"04052NF",
    "Geometry Common Core":"02072",
    "Geometry H Common Core":"02072",
    "English 10 H":"01003CC",
    "English 11":"01003CC",
    "Earth and Space Science":"03001",
    "Spanish II H":"NONE",
    "Spanish III":"NONE"
    }

dict5 = {
    "StudentID":stuIDList,
    "LastName": lNameList,
    "FirstName": fNameList,
    "StateCourseCode": courseList,
    "CourseSection": sectionList,
    "TeacherName": teacherList,
    "GradeLevel": gradeList
    }

codeDict = {}

dict7 = {
    "Host ID":hostIDList,
    "Course title": courseList,
    "Numeric grade": scoreList,
    "School year": junkList,
    "First name": junkList,
    "Last name": junkList,
    "Grade plan grade": junkList,
    "Head teacher": junkList,
    "Home phone": junkList,
    "Graduation year": junkList,
    "Parents/Guardians": junkList,
    "Gender": junkList,
    "Departments": junkList,
    "Added on": junkList
    }

dict8 = { # RegentsUpdated test description coordinator
    "Algebra 1 Common Core":"Algebra I Common Core Regents",
    "Algebra 1H Common Core":"Algebra I Common Core Regents",
    "Algebra II":"Algebra II Common Core Regents",
    "Algebra II Honors":"Algebra II Common Core Regents",
    "Chemistry":"Chemistry Regents",
    "U.S. History":"US History and Government (Framework) Regents",
    "U.S. History H":"US History and Government (Framework) Regents",
    "AP U.S. History":"US History and Government (Framework) Regents",
    "Physics Honors":"Physics Regents",
    "Living Envir Bio":"Living Environment - Biology Regents",
    "Living Envir-Bio H":"Living Environment - Biology Regents",
    "Global History 10":"Global History & Geography II Regents",
    "Global History 10 H":"Global History & Geography II Regents",
    "AP World History: Modern":"Global History & Geography II Regents",
    "Geometry Common Core":"Geometry Common Core Regents",
    "Geometry H Common Core":"Geometry Common Core Regents",
    "English 10 H":"English Regents",
    "English 11":"English Regents",
    "Earth Science":"Earth Science Regents",
    "Spanish III":"FLACS B: Spanish",
    "Spanish III H":"FLACS B: Spanish"
    }

dict9 = {
    "STU_ID":stuIDList,
    "FNAME":fNameList,
    "LNAME": lNameList,
    "MI": mNameList,
    "GRADE": gradeList,
    "ADD1": addressList,
    "ADD2": address2List,
    "CITY": cityList,
    "STATE": stateList,
    "ZIP": zipCodeList,
    "GEN": genderList,
    "DOB": dobList,
    "Parent 1 First Name":p1fnList,
    "Parent 1 Last Name":p1lnList,
    "Parent 2 First Name":p2fnList,
    "Parent 2 Last Name":p2lnList,
    "HR":hrList,
    "Parent 1 Email": p1emailList,
    "Parent 1 Cell": p1cellList,
    "Parent 1 Work": p1workList,
    "Parent 1 Home": p1homeList,
    "Parent 2 Work": p2workList,
    "I20": I20List,
    "Student Email": stuEmailList,
    "District": districtList}
    

#print(dict1)
def demographicUpload():
    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    filename = askopenfilename(title= "Please Select the Demographic Info File from MSA:", \
                               filetypes = [("Excel Files", "*.xlsx")]) # show an "Open" dialog box and return the path to the selected file
    filename2 = askopenfilename(title= "Enter District Code Spreadsheet Here:", \
                                filetypes = [("Excel Files", "*.xlsx")])
    #print(filename)
    dataframe = openpyxl.load_workbook(filename) # openpyxl stuff to read from initial spreadsheet
    dataframe1 = dataframe.active
    wb = openpyxl.Workbook()
    sheet = wb.active
    dataframe2 = openpyxl.load_workbook(filename2)
    dataframe3 = dataframe2.active
    wb3 = openpyxl.Workbook()
    sheet3 = wb3.active

    """
    for row in dataframe1:
        if row[0].value == None:
            name = ""
        else:
            name = row[0].value
        nameList.append(name)

    for row in dataframe1:
        food = row[1].value
        foodList.append(food)
    """
    # Sucks in the data from the spreadsheet you choose and assigns it to the proper list.
    for col in range (0, dataframe1.max_column):
        #print(col)
        for row in dataframe1.iter_rows(1, dataframe1.max_row):
            try:
                x = dict1[str(row[col].value)]
            except:
                pass
            #print(row[col].value)
            if row[col].value == None:
                x.append("")
            else:
                x.append(row[col].value)
            #print(x)

    #Adding district codes to a dictionary
    for row in range(1, dataframe3.max_row + 1):
        key = dataframe3.cell(row, 1).value
        value = dataframe3.cell(row, 2).value
        codeDict[key] = value
          
    # Where all the organizing happens. Some shortened, some formatted, just string manipulation.

    for i in range(0,len(gradeList)):
        longCodeList.append(longCode)
        shortCodeList.append(shortCode)
        regentDateList.append(regentDate)
        aList.append(A)
        spacerList.append("")
        try:
            if zipCodeList[i+1] not in codeDict:
                codeDict[zipCodeList[i+1]] = "NO DISTRICT CODE FOUND FOR ZIPCODE"
            z = codeDict[zipCodeList[i+1]]
            distList.append(str(z))
            y = dict3[gradeList[i+1]]
            startDate = str(y) + "-09-01"
            startList.append(startDate)
            if gradeList[i+1] == 9:
                gradeList[i+1] = "0"+str(gradeList[i+1])
            if mNameList[i+1] == "":
                pass
            else:
                mNameList[i+1] = mNameList[i+1][0]
            if genderList[i+1] == "":
                pass
            else:
                genderList[i+1] = genderList[i+1][0]
            if hispanicList[i+1] == "":
                pass
            else:
                hispanicList[i+1] = hispanicList[i+1][0]
            if dobList[i+1] == "":
                pass
            else:
                dobList[i+1] = dobList[i+1].strftime("%Y-%m-%d")
            p1fullList.append(p1fnList[i+1] + " " + p1lnList[i+1])
            p2fullList.append(p2fnList[i+1] + " " + p2lnList[i+1])
            if ethnicityList[i+1] == "":
                pass
            else:
                ethnicityList[i+1] = dict2[ethnicityList[i+1][0]]
            hrList[i+1] = hrList[i+1].replace("Homeroom ", "")
        except:
            pass

    # New spreadsheet to write into

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active

    #Writes all of the lists into columns

    for row in zip(longCodeList, shortCodeList, regentDateList, stuIDList, \
                   lNameList, fNameList, mNameList, gradeList, hrList, dobList, \
                   genderList, ethnicityList, spacerList, spacerList, \
                   spacerList, spacerList, spacerList, spacerList, \
                   aList, regentDateList, spacerList, spacerList, \
                   spacerList, spacerList, spacerList, startList, \
                   spacerList, spacerList, addressList, \
                   address2List, cityList, stateList, zipCodeList, \
                   p1fullList, p2fullList, spacerList, spacerList, \
                   spacerList, spacerList, spacerList, distList, hispanicList, \
                   ):
        sheet2.append(row)

    for k in range(1, 52):
        if sheet2.cell(1,k).value == "" or sheet2.cell(1,k).value == None:
              sheet2.cell(1,k).value = k

    """
    for row in range(0, dataframe1.max_row):
            for col in dataframe1.iter_cols(1,dataframe1.max_column):
                if (col[row].value == "Name"):
                    print(col[row].value)
    """

    #Ask user where they want to save the file
    folder_selected = filedialog.askdirectory(title="Please Select Where You Want To Save Your New File")
    folder_selected = folder_selected +"\\demographicInformation.xlsx"
    print(folder_selected.replace("/","\\"))
    wb2.save(folder_selected)
    sys.exit()

def regentsAnswerSheets():
    Tk().withdraw()
    filename = askopenfilename(title= "Please Select the RIC Answer Sheet File from MSA:", \
                               filetypes = [("Excel Files", "*.xlsx")])
    dataframe = openpyxl.load_workbook(filename) # openpyxl stuff to read from initial spreadsheet
    dataframe1 = dataframe.active
    # Sucks in data
    for col in range (0, dataframe1.max_column):
        #print(col)
        for row in dataframe1.iter_rows(1, dataframe1.max_row):
            try:
                x = dict5[str(row[col].value)]
            except:
                pass
            #print(row[col].value)
            if row[col].value == None:
                x.append("")
            else:
                x.append(row[col].value)
           #print(x)
    #print(len(courseList))
    #print(len(sectionList))
    #Replaces Course name with corresponding regents Code
    for i in range (0, len(courseList)):
        try:
            courseList[i+1] = dict4[courseList[i+1]]
        except:
            pass
    #print(courseList)
    # Setup for some lists
    for i in range(0,len(courseList)-1):
        if gradeList[i] == 9:
            gradeList[i] = "0" + str(gradeList[i])
        stuIDList[i] = "0000"+str(stuIDList[i])
        longCodeList.append(longCode)
        shortCodeList.append(shortCode)
        schoolList.append(school)
        spacerList.append("")
        regentDateList.append(regentDate)
        adminMonthList.append(textMonth)
    # Creating the sheets for different regent subjects
    wb = openpyxl.Workbook()
    algebra1 = wb.active
    wb_sheet = wb['Sheet']
    wb_sheet.title = ("algebra1")
    algebra2 = wb.create_sheet("algebra2")
    physics = wb.create_sheet("physics")
    chemistry = wb.create_sheet("chemistry")
    usHistory = wb.create_sheet("usHistory")
    livingEnvir = wb.create_sheet("livingEnvir")
    globalHistory = wb.create_sheet("globalHistory")
    geometry = wb.create_sheet("geometry")
    english = wb.create_sheet("english")
    earthScience = wb.create_sheet("earthScience")
    spanish = wb.create_sheet("spanish")
    #linking regent code to sheet
    dict6 = {
    "02050": algebra1,
    "02056CC": algebra2,
    "03101": chemistry,
    "04101F": usHistory,
    "03151": physics,
    "03051": livingEnvir,
    "04052NF": globalHistory,
    "02072": geometry,
    "01003CC": english,
    "03001": earthScience,
    "NONE": spanish
    }
    #print(wb.sheetnames)
    #Sheet Titles
    for sheet in wb:
        sheet.append(titleList)
    #Put data where it belongs
    for row in zip(longCodeList, shortCodeList, regentDateList, adminMonthList, stuIDList, lNameList, \
                   fNameList, gradeList, sectionList, teacherList, courseList, \
                   spacerList, schoolList):
        try:
            dict6[row[10]].append(row)
        except:
            pass
    #print(row[1].value)
    #print(dict6[courseList[1]])
    #print(wb[dict6[courseList[row[0].value]]])
    # Save File
    folder_selected = filedialog.askdirectory(title="Please Select Where You Want To Save Your New File")
    folder_selected = folder_selected +"\\RegentsWorkbook.xlsx"
    print(folder_selected.replace("/","\\"))
    wb.save(folder_selected)
    sys.exit()

def regentsUpdater():
    stuIDList.append("student_id")
    subtest_descriptionList = []
    Tk().withdraw()
    filename = askopenfilename(title= "Please Select the Initial Regents Grade Export File from MSA:", \
                               filetypes = [("Excel Files", "*.xlsx")])
    
    dataframe = openpyxl.load_workbook(filename) # openpyxl stuff to read from initial spreadsheet
    dataframe1 = dataframe.active
    for col in range (0, dataframe1.max_column):
        #print(col)
        for row in dataframe1.iter_rows(1, dataframe1.max_row):
            try:
                x = dict7[str(row[col].value)]
            except:
                pass
            #print(row[col].value)
            if row[col].value == None:
                x.append("")
            else:
                x.append(row[col].value)
    for i in range (0, len(courseList)):
        try:
            courseList[i+1] = dict8[courseList[i+1]]
        except:
            pass
    for j in range(0, len(courseList)):
        subtest_descriptionList.append(courseList[j])
        subtest_descriptionList[j] = subtest_descriptionList[j].replace(" Regents","")
        subtest_descriptionList[j] = subtest_descriptionList[j].replace("FLACS B:","")
        subtest_descriptionList[j] = subtest_descriptionList[j].replace(" - Biology","")
    for k in range(0, len(courseList)-1):
        stuIDList.append("")
        userID.append("")
        testDateTime.append("6/30/" + str(year) + "12:00:00 AM")
        comment.append("")
        overallScore.append("")
        registeredDate.append("")
        registeredLocation.append("")
        rescheduled.append("")
        noShow.append("")
        proctorID.append("")
        proctorOther.append("")
        printOnReportCard.append("")
        printOnTranscript.append("Yes")
        percentile.append("")
        scale.append("")
        stanine.append("")
        subtestComment.append("")
        

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active

    subtest_descriptionList[0] = "subtest_description"
    hostIDList[0] = "host_id"
    courseList[0] = "test_description"
    scoreList[0] = "score"

    for row in zip(stuIDList, hostIDList, userID, courseList, testDateTime, comment, overallScore, registeredDate, \
                   registeredLocation, rescheduled, noShow, proctorID, proctorOther, printOnReportCard, \
                   printOnTranscript, subtest_descriptionList, scoreList, percentile, scale, stanine, subtestComment):
        sheet2.append(row)

    folder_selected = filedialog.askdirectory(title="Please Select Where You Want To Save Your New File")
    folder_selected = folder_selected +"\\RegentsScores.xlsx"
    print(folder_selected.replace("/","\\"))
    wb2.save(folder_selected) 
    sys.exit()

def manhassetDistrictReporting():
    filename = askopenfilename(title= "Please Select the Manhasset District File from MSA:", \
                               filetypes = [("Excel Files", "*.xlsx")]) # show an "Open" dialog box and return the path to the selected file
    filename2 = askopenfilename(title= "Enter District Code Spreadsheet Here:", \
                                filetypes = [("Excel Files", "*.xlsx")])

    dataframe = openpyxl.load_workbook(filename) # openpyxl stuff to read from initial spreadsheet
    dataframe1 = dataframe.active
    wb = openpyxl.Workbook()
    sheet = wb.active
    dataframe2 = openpyxl.load_workbook(filename2)
    dataframe3 = dataframe2.active
    wb3 = openpyxl.Workbook()
    sheet3 = wb3.active

    # Sucks in the data from the spreadsheet you choose and assigns it to the proper list.
    for col in range (0, dataframe1.max_column):
        #print(col)
        for row in dataframe1.iter_rows(1, dataframe1.max_row):
            try:
                x = dict9[str(row[col].value)]
            except:
                pass
            #print(row[col].value)
            if row[col].value == None:
                x.append("")
            else:
                x.append(row[col].value)
            #print(x)

    #Adding district codes to a dictionary
    for row in range(1, dataframe3.max_row + 1):
        key = dataframe3.cell(row, 1).value
        value = dataframe3.cell(row, 2).value
        codeDict[key] = value

    for i in range(0,len(stuIDList)):
        try:
            if dobList[i+1] == "":
                pass
            else:
                dobList[i+1] = dobList[i+1].strftime("%m/%d/%Y")
            gradeList[i+1] = str(gradeList[i+1]) + "th"
            if p1workList[i+1] == "":
                p1workList[i+1] = p2workList[i+1]
            if zipCodeList[i+1] not in codeDict:
                codeDict[zipCodeList[i+1]] = "NO DISTRICT CODE FOUND FOR ZIPCODE"
            z = codeDict[zipCodeList[i+1]]
            distList.append(str(z))
            dobList[i+1] = str(dobList[i+1])
        except:
            pass

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active

    #print(dobList)

    for row in zip(stuIDList, fNameList, lNameList, mNameList, gradeList, hrList, p1fnList, p1lnList, \
                   p2fnList, p2lnList, p1cellList, p1workList, p1homeList, p1emailList, addressList, \
                   address2List, cityList, stateList, zipCodeList, genderList, districtList, distList, \
                   stuEmailList, dobList, I20List):
        sheet2.append(row)

    folder_selected = filedialog.askdirectory(title="Please Select Where You Want To Save Your New File")
    folder_selected = folder_selected +"\\ManhassetDistrictUpload.xlsx"
    print(folder_selected.replace("/","\\"))
    wb2.save(folder_selected) 
    sys.exit()


# Runner Function
def runCode():
    global entry
    option = entry.get()
    if option == "1":
        #print("Option 1 Selected")
        win.destroy()
        demographicUpload()
    elif option == "2":
        #print("Option 2 Selected")
        win.destroy()
        regentsAnswerSheets()
    elif option == "3":
        win.destroy()
        regentsUpdater()
    elif option == "4":
        win.destroy()
        manhassetDistrictReporting()
    else:
        label.config(text= "Please choose one of the options below")

# Small GUI
win = Tk()
win.title("Saint Mary's Spreadsheet Organizer")
win.geometry("400x300")
label=Label(win, text="Welcome to the Saint Mary's Spreadsheet Organizer\n \
    Please enter a choice according to your needs.")
label2=Label(win, text= "1: Demographic Upload\n2: Regents Answer Sheets\n\
3: Regents Scores Updater\n4: Manhasset District Reporting\n")
label.pack()
label2.pack()
entry = Entry(win, width=20)
entry.focus_set()
entry.pack()

ttk.Button(win, text= "Click to Start", width = 20, command = \
           runCode).pack(pady=20)
           
win.mainloop()           


